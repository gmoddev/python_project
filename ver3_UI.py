import zipfile
import os
import openpyxl
import tkinter as tk
from tkinter import filedialog, Entry
from PIL import Image, ImageTk
from datetime import datetime

def count_file_types(zip_ref, folder_path, descendants_count, visited_folders):
    if folder_path in visited_folders:
        return

    visited_folders.add(folder_path)

    for file_info in zip_ref.infolist():
        file_path = file_info.filename
        if file_path.startswith(folder_path):
            is_directory = file_path.endswith('/')
            file_type = 'folder' if is_directory else file_path.split('.')[-1].lower()
            modified_date = datetime(*file_info.date_time).strftime('%Y-%m-%d %H:%M:%S')
            descendants_count[file_path] = f'Type: {file_type}, Modified Date: {modified_date}'

            if is_directory:
                count_file_types(zip_ref, file_path, descendants_count, visited_folders)

def save_hierarchy_to_file(file_path, hierarchy):
    with open(file_path, 'w') as file:
        for path, info in hierarchy.items():
            file.write(f"{path} - {info}\n")

def read_hierarchy_from_file(file_path):
    hierarchy = {}
    with open(file_path, 'r') as file:
        for line in file:
            parts = line.strip().split(' - ')
            if len(parts) == 2:
                path, info = parts
                info_parts = info.split(', ')
                if len(info_parts) == 2:
                    file_type, modified_date = info_parts
                    hierarchy[path.strip()] = {'type': file_type, 'modified_date': modified_date.strip()}
    return hierarchy

def create_excel_spreadsheet(output_excel, hierarchy):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='File Path')
    ws.cell(row=1, column=2, value='File Type')
    ws.cell(row=1, column=3, value='Modified Date')

    for row, (path, info) in enumerate(hierarchy.items(), start=2):
        file_type = info['type']
        modified_date = info['modified_date']
        ws.cell(row=row, column=1, value=path)
        ws.cell(row=row, column=2, value=file_type)
        ws.cell(row=row, column=3, value=modified_date)

    wb.save(output_excel)
    print(f"Excel spreadsheet saved to {output_excel}")

def convert_excel_to_text(excel_file, output_text_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        with open(output_text_file, 'w') as text_file:
            for row in ws.iter_rows(values_only=True):
                line = ' - '.join(str(cell) for cell in row)
                text_file.write(line + '\n')

        print(f"Text file '{output_text_file}' created successfully.")

    except Exception as e:
        print(f"Error: {e}")

class SpreadsheetApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Spreadsheet Tool")
        self.geometry("315x275")
        self.minsize(315, 275)

        icon_path = "./aiden.ico"
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)

        self.bg_image_path = "./aiden.png"
        self.create_ui()
        self.bind("<Configure>", self.update_geometry_label)

    def load_background_image(self, bg_image_path, width, height):
        original_bg_image = Image.open(bg_image_path)
        resized_bg_image = original_bg_image.resize((width, height), Image.BICUBIC)
        bg_image = ImageTk.PhotoImage(resized_bg_image)
        return bg_image
    
    def create_ui(self):
        main_frame = tk.Frame(self, bg="", bd=0)  # Set background to empty string for transparency
        main_frame.pack(fill=tk.BOTH, expand=True)

        if os.path.exists(self.bg_image_path):
            original_bg_image = Image.open(self.bg_image_path)
            self.bg_image = ImageTk.PhotoImage(original_bg_image)

            bg_label = tk.Label(main_frame, image=self.bg_image)
            bg_label.image = self.bg_image
            bg_label.place(relwidth=1, relheight=1)

        self.entry_zip = Entry(main_frame, width=50, bg="white")
        self.entry_zip.insert(0, "Path to Zip File")
        self.entry_zip.pack(pady=5)

        btn_hierarchy = tk.Button(main_frame, text="Generate Hierarchy", command=self.generate_hierarchy, bg="#4caf50", fg="white", font=("Helvetica", 12))
        btn_hierarchy.pack(pady=10)

        self.entry_excel = Entry(main_frame, width=50, bg="white")
        self.entry_excel.insert(0, "Path to Excel File")
        self.entry_excel.pack(pady=5)

        btn_convert = tk.Button(main_frame, text="Convert Excel to Text", command=self.convert_excel_to_text, bg="#4caf50", fg="white", font=("Helvetica", 12))
        btn_convert.pack(pady=10)

        self.entry_hierarchy = Entry(main_frame, width=50, bg="white")
        self.entry_hierarchy.insert(0, "Path to Hierarchy Text File")
        self.entry_hierarchy.pack(pady=5)

        btn_spreadsheet = tk.Button(main_frame, text="Create Spreadsheet", command=self.create_spreadsheet, bg="#4caf50", fg="white", font=("Helvetica", 12))
        btn_spreadsheet.pack(pady=10)

        self.geometry_label = tk.Label(main_frame, text=f"Current Geometry: {self.geometry()}", bg="SystemButtonFace", relief=tk.FLAT)
        self.geometry_label.pack()

        self.footer_label = tk.Label(main_frame, text="Made by not_lowest", font=("Helvetica", 8), pady=5, bg="SystemButtonFace", relief=tk.FLAT)
        self.footer_label.pack()

    def get_entry_value(self, entry):
        return entry.get().strip()

    def update_geometry_label(self, event):
        self.geometry_label.config(text=f"Current Geometry: {self.geometry()}")
        new_width = self.winfo_width()
        new_height = self.winfo_height()
        self.bg_image = self.load_background_image(self.bg_image_path, new_width, new_height)
        bg_label = self.children["!frame"].children["!label"]
        bg_label.configure(image=self.bg_image)

    def generate_hierarchy(self):
        zip_file_path = self.get_entry_value(self.entry_zip)
        if zip_file_path:
            output_file_path = filedialog.asksaveasfilename(title="Save Hierarchy As", defaultextension=".txt",
                                                             filetypes=[("Text Files", "*.txt")])
            if output_file_path:
                with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                    descendants_count = {}
                    visited_folders = set()
                    count_file_types(zip_ref, '', descendants_count, visited_folders)

                save_hierarchy_to_file(output_file_path, descendants_count)
                print(f"Hierarchy information saved to {output_file_path}")

    def convert_excel_to_text(self):
        excel_file_path = self.get_entry_value(self.entry_excel)
        if excel_file_path:
            output_text_file_path = filedialog.asksaveasfilename(title="Save Text File As", defaultextension=".txt",
                                                                  filetypes=[("Text Files", "*.txt")])
            if output_text_file_path:
                convert_excel_to_text(excel_file_path, output_text_file_path)

    def create_spreadsheet(self):
        hierarchy_file_path = self.get_entry_value(self.entry_hierarchy)
        if hierarchy_file_path:
            output_excel_path = filedialog.asksaveasfilename(title="Save Excel As", defaultextension=".xlsx",
                                                              filetypes=[("Excel Files", "*.xlsx")])
            if output_excel_path:
                hierarchy = read_hierarchy_from_file(hierarchy_file_path)
                create_excel_spreadsheet(output_excel_path, hierarchy)
                print(f"Excel spreadsheet saved to {output_excel_path}")

if __name__ == "__main__":
    app = SpreadsheetApp()
    app.mainloop()
