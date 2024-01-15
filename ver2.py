import zipfile
import os
import openpyxl

from tabulate import tabulate
from datetime import datetime

def count_file_types(zip_ref, folder_path, descendants_count, visited_folders):
    if folder_path in visited_folders:
        return

    visited_folders.add(folder_path)

    for file_info in zip_ref.infolist():
        # Extract file path from the zip folder
        file_path = file_info.filename

        # If the file is within the specified folder
        if file_path.startswith(folder_path):
            # Determine file type (folder, xls, file, etc.)
            is_directory = file_path.endswith('/')
            file_type = 'folder' if is_directory else file_path.split('.')[-1].lower()

            # Get the modified date of the file
            modified_date = datetime(*file_info.date_time).strftime('%Y-%m-%d %H:%M:%S')

            # Update the count in the dictionary
            descendants_count[file_path] = f'Type: {file_type}, Modified Date: {modified_date}'

            # If it's a folder, recursively count descendants
            if is_directory:
                count_file_types(zip_ref, file_path, descendants_count, visited_folders)



### FILE DATA FUNCS

def save_hierarchy_to_file(file_path, hierarchy):
    with open(file_path, 'w') as file:
        for path, info in hierarchy.items():
            file.write(f"{path} - {info}\n")

def read_hierarchy_from_file(file_path):
    hierarchy = {}
    with open(file_path, 'r') as file:
        for line in file:
            parts = line.strip().split(' - ')
            if len(parts) == 2:
                path, info = parts
                info_parts = info.split(', ')
                if len(info_parts) == 2:
                    file_type, modified_date = info_parts
                    hierarchy[path.strip()] = {'type': file_type, 'modified_date': modified_date.strip()}
    return hierarchy

## Excel Funcs

def convert_excel_to_text(excel_file, output_text_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        with open(output_text_file, 'w') as text_file:
            for row in ws.iter_rows(values_only=True):
                line = ' - '.join(str(cell) for cell in row)
                text_file.write(line + '\n')

        print(f"Text file '{output_text_file}' created successfully.")

    except Exception as e:
        print(f"Error: {e}")

def create_excel_spreadsheet(output_excel, hierarchy):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hierarchy'

    # Add headers
    ws.cell(row=1, column=1, value='File Path')
    ws.cell(row=1, column=2, value='File Type')
    ws.cell(row=1, column=3, value='Modified Date')

    # Add data
    for row, (path, info) in enumerate(hierarchy.items(), start=2):
        file_type = info['type']
        modified_date = info['modified_date']
        ws.cell(row=row, column=1, value=path)
        ws.cell(row=row, column=2, value=file_type)
        ws.cell(row=row, column=3, value=modified_date)

    wb.save(output_excel)
    print(f"Excel spreadsheet saved to {output_excel}")

## MAIN FUNCTION

def main():
    def_path = "C:/Users/aiden/Downloads"
    zip_path = os.path.join(def_path, 'WindowsProject(3912).zip')
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    output_file = os.path.join(def_path, f'hierarchy_output_{timestamp}.txt')
    output_excel = f'hierarchy_spreadsheet_{timestamp}.xlsx'

    if os.path.exists(zip_path):
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            descendants_count = {}
            visited_folders = set()
            count_file_types(zip_ref, '', descendants_count, visited_folders)

        # Save the hierarchy to a text file
        save_hierarchy_to_file(output_file, descendants_count)

        hierarchy = read_hierarchy_from_file(output_file)
        
        create_excel_spreadsheet(output_excel, hierarchy)

        # Print a message
        print(f"Hierarchy information saved to {output_file}")

    else:
        print(f"Error: The specified path '{zip_path}' does not exist.")

if __name__ == "__main__":
    main()
